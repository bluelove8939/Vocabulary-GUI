import os
import random

import openpyxl as op
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, NamedStyle


os_user_name = os.getlogin()
document_path = os.path.join("C:/Users", os_user_name, "Documents")
current_dirname = os.path.split(os.path.abspath(__file__))[0].strip()
default_dir = os.path.join("C:/Users", os_user_name, "Documents/Vocabulary sheets")
default_name = 'problem file'
default_ftype = '*.xlsx'
default_font = 'Arial'
supporting_types = [('Basic Excel 2007 files (*.xlsx)', '*.xlsx'), ('Excel 2007 files (*.xlsm)', '*.xlsm'), ('Excel 2007 formats (*.xltx)', '*.xltx'), ('Excel 2007 templates (*.xltm)', '*.xltm')]
testfiles = [os.path.join(current_dirname, file) for file in ["test/vocab sheet 1.xlsx", "test/vocab sheet 2.xlsx"]]
ftypes = [item[1] for item in supporting_types]

if "Vocabulary sheets" not in os.listdir(document_path): os.mkdir(default_dir)


# FileGenerator class
class FileGenerator:
    def __init__(self):
        self.queries = []
        self.files = []
        self.answers = {}

        # Style parameters
        self.bd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        self.ft = Font(name=default_font, size=15, bold=False)
        self.bold_ft = Font(name=default_font, size=15, bold=True)
        self.al = Alignment(vertical='center', horizontal='left')

    def clear(self):
        self.queries = []
        self.files = []
        self.answers = {}

    def set_font(self, name=default_font, size=15):
        self.ft = Font(name=name, size=size, bold=False)
        self.bold_ft = Font(name=name, size=size, bold=True)

    def read(self, filename, que_pivot=0, ans_pivot=1, index=True):
        currentfile = op.load_workbook(filename)
        self.files.append(filename)

        for sheet in currentfile:
            for idx, row in enumerate(sheet.values):
                if not index or idx != 0:
                    self.queries.append(row[que_pivot])
                    self.answers[row[que_pivot]] = row[ans_pivot]

    def make_problems(self, dirname=default_dir, name=default_name, ftype=default_ftype, size=None, random_seed=None):
        # Define filename
        name = FileGenerator.DEFAULT_NAME(name, dirname)
        filename = os.path.join(dirname, name + ftype[1:])

        # Shuffle query list
        random.seed(a=random_seed)
        random.shuffle(self.queries)

        queries = self.queries[:size if size is None else len(self.queries)]

        # Make newfile
        newfile = Workbook()

        newfile.active.cell(row=1, column=2).value = "Problems"
        newfile.active.cell(row=1, column=3).value = "Answers"

        for idx, query in enumerate(queries):
            newfile.active.cell(row=idx+2, column=1).value = idx+1
            newfile.active.cell(row=idx+2, column=2).value = query

        # Style assignment
        base_style = NamedStyle(name='base_problem', alignment=self.al, font=self.ft, border=self.bd)
        index_style = NamedStyle(name='index_problem', alignment=self.al, font=self.bold_ft, border=self.bd)

        for ridx, row in enumerate(newfile.active):
            for cidx, cell in enumerate(row):
                if ridx == 0 or cidx == 0: cell.style = index_style
                else: cell.style = base_style
        
        # Cell width and height assignment
        side_col_width = 0
        query_col_width = 50
        answer_col_width = 50

        for idx, row in enumerate(newfile.active.values):
            side_col_width = max(side_col_width, len(str(row[0]))*1.2)
            query_col_width = max(query_col_width, len(row[1])*1.5)
            newfile.active.row_dimensions[idx+1].height = 30

        newfile.active.column_dimensions['A'].width = side_col_width
        newfile.active.column_dimensions['B'].width = query_col_width
        newfile.active.column_dimensions['C'].width = answer_col_width

        newfile.save(filename)

        return filename

    def make_answers(self, prob_filename, dirname="DEFAULT_DIR", name="DEFAULT_NAME", ftype="DEFAULT_FTYPE"):
        # Define filename
        prob_dirname, prob_name = os.path.split(prob_filename)
        prob_name, prob_ftype = prob_name.split('.')
        prob_ftype = '.' + prob_ftype

        if dirname == "DEFAULT_DIR": dirname = prob_dirname
        if name == "DEFAULT_NAME":
            if ftype == "DEFAULT_FTYPE": ftype = prob_ftype
            name = FileGenerator.DEFAULT_NAME("[Answer] " + prob_name) + ftype
        filename = os.path.join(dirname, name)
        
        # Make newfile
        newfile = op.load_workbook(prob_filename)

        for idx, row in enumerate(newfile.active):
            if idx != 0 and row[1].value in self.answers.keys():
                row[2].value = self.answers[row[1].value]

        # Style assignment
        base_style = NamedStyle(name='base_answer', alignment=self.al, font=self.ft, border=self.bd)
        index_style = NamedStyle(name='index_answer', alignment=self.al, font=self.bold_ft, border=self.bd)

        for ridx, row in enumerate(newfile.active):
            for cidx, cell in enumerate(row):
                if ridx == 0 or cidx == 0: cell.style = index_style
                else: cell.style = base_style

        newfile.save(filename)

        return filename

    def make_both(self, dirname=default_dir, name=default_name, ftype=default_ftype, size=None, random_seed=None):
        prob_filename = self.make_problems(dirname, name, ftype, size, random_seed)
        ans_filename = self.make_answers(prob_filename)

    @staticmethod
    def DEFAULT_NAME(default, dirname=default_dir):
        cnt = 0
        for name in os.listdir(dirname):
            if default == name[:len(default)]:
                cnt += 1
    
        if cnt == 0: return default
    
        return default + ' ' + str(cnt+1)


if __name__ == "__main__":
    nf = FileGenerator()
    for file in testfiles: nf.read(file)
    prob = nf.make_problems()
    ans = nf.make_answers(prob_filename=prob)